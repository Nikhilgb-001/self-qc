import streamlit as st
import pandas as pd
from docx import Document
from io import BytesIO
import re
from openpyxl import load_workbook

# --- SETTINGS ---
st.set_page_config(page_title="Automated Agreement Validation System", layout="centered")

hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)



# Normalize function
def normalize(text):
    if text is None:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip().lower().replace("\u200b", "").replace("\xa0", " ")

# Extract fields from Word table
def extract_fields_from_word(file):
    doc = Document(file)
    extracted = {}
    for table in doc.tables:
        for row in table.rows:
            if len(row.cells) >= 2:
                field = row.cells[0].text.strip()
                value = row.cells[1].text.strip()
                if field:
                    extracted[normalize(field)] = value
    return extracted

# --- UI START ---

# Company Logo
st.image('app/teva.png', width=150)

st.markdown(
    """
    <h2 style='text-align: center;'>Automated Agreement Validation System</h2>
    <p style='text-align: center;'>Upload your finalized Agreement Document (.docx) and Compliance Checklist (.xlsx) for automated field validation.</p>
    """, unsafe_allow_html=True
)
st.divider()

# Upload Files
st.subheader("1️⃣ Upload Files")

# Upload columns side-by-side
col1, col2 = st.columns(2)

with col1:
    st.image('app/word.jpeg', width=50)
    docx_file = st.file_uploader("Upload Agreement Document (.docx)", type=["docx"], key="word_upload")

with col2:
    st.image('app/Excel.jpeg', width=50)
    excel_file = st.file_uploader("Upload Agreement Masterfile (.xlsx)", type=["xlsx"], key="excel_upload")

# Processing logic
if docx_file and excel_file:
    with st.spinner('🔎 Extracting and Matching Fields... Please wait...'):
        word_data = extract_fields_from_word(docx_file)

        output = BytesIO()
        output.write(excel_file.read())
        output.seek(0)
        workbook = load_workbook(output)

        preview_rows = []

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            headers = [cell.value for cell in ws[1]]
            clean_headers = [normalize(h) for h in headers]

            field_col_idx = None
            value_col_idx = None

            for idx, header in enumerate(clean_headers):
                if "fieldname" in header or header == "field":
                    field_col_idx = idx + 1
                elif "value" in header:
                    value_col_idx = idx + 1

            if field_col_idx and value_col_idx:
                for row in range(2, ws.max_row + 1):
                    field_cell = ws.cell(row=row, column=field_col_idx)
                    manual_cell = ws.cell(row=row, column=value_col_idx)

                    if field_cell.value:
                        normalized_field_excel = normalize(field_cell.value)

                        if normalized_field_excel in word_data:
                            preview_rows.append({
                                "Sheet Name": sheet_name,
                                "Field Name": field_cell.value,
                                "Old Manual Value": manual_cell.value if manual_cell.value else "",
                                "New Extracted Value": word_data[normalized_field_excel]
                            })

    st.divider()

    # Preview Section
    st.subheader("2️⃣ Review Matching Fields")

    if preview_rows:
        st.success(f"✅ Found {len(preview_rows)} matching fields!")
        preview_df = pd.DataFrame(preview_rows)
        st.dataframe(preview_df, use_container_width=True)

        st.divider()
        st.subheader("3️⃣ Start Validation & Download")

        if st.button("🚀 Start Review and Download Updated Checklist"):
            with st.spinner('⚡ Applying Updates...'):
                for match in preview_rows:
                    ws = workbook[match["Sheet Name"]]
                    headers = [cell.value for cell in ws[1]]
                    clean_headers = [normalize(h) for h in headers]

                    field_col_idx = None
                    value_col_idx = None

                    for idx, header in enumerate(clean_headers):
                        if "fieldname" in header or header == "field":
                            field_col_idx = idx + 1
                        elif "value" in header:
                            value_col_idx = idx + 1

                    for row in range(2, ws.max_row + 1):
                        field_cell = ws.cell(row=row, column=field_col_idx)
                        if field_cell.value and normalize(field_cell.value) == normalize(match["Field Name"]):
                            manual_cell = ws.cell(row=row, column=value_col_idx)
                            manual_cell.value = match["New Extracted Value"]
                            break

                final_output = BytesIO()
                workbook.save(final_output)
                final_output.seek(0)

                st.success("🎯 Validation Complete! Download your updated file below:")

                st.download_button(
                    label="📥 Download Updated Masterfile",
                    data=final_output.getvalue(),
                    file_name="Updated_Checklist.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.warning("⚠️ No matching fields found between Agreement Document and Masterfile.")
else:
    st.info("📂 Please upload both documents to start validation.")

st.divider()

# Footer Note
st.caption("This tool is intended for internal use within Teva Pharmaceuticals' Compliance Team only. Unauthorized disclosure of uploaded documents strictly prohibited.")

